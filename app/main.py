from typing import Optional, List, Dict
from pathlib import Path
import csv
import io
from collections import defaultdict

from fastapi import FastAPI, Request, Form, Query, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from sqlmodel import SQLModel, Session, create_engine, select
from sqlalchemy import func, text
from sqlalchemy.exc import OperationalError
import logging
from datetime import datetime
from types import SimpleNamespace
import time
import hmac
import hashlib

from starlette.middleware.sessions import SessionMiddleware
from passlib.context import CryptContext

# models.py から定義をインポート
from .models import (
    Article,
    ScreeningDecision,
    User,
    AppConfig,
    ScaleArticle,
    ScaleScreeningDecision,
    SecondaryArticle,
    SecondaryAutoExtraction,
    SecondaryReview,
)
import os

# Load environment from dotenv: allow explicit ENV_FILE or fallback to local
from dotenv import load_dotenv

# 明示指定があればそれを読む（例: ENV_FILE=/path/to/.env.prod）
env_file = os.getenv("ENV_FILE")
if env_file:
    load_dotenv(env_file)
else:
    # ローカル開発の標準
    load_dotenv(".env.local")

# =========================================================
# 基本設定
# =========================================================
# Application base directory
BASE_DIR = Path(__file__).resolve().parent.parent
# DATABASE_URL must be provided by the environment. Do NOT fall back to a hardcoded path.
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL environment variable is required and must point to the main DB")

N_GROUPS = 4

# グループ名の定義
GROUP_NAMES = {
    1: "Jyunten",
    2: "Osaka",
    3: "Nagoya1",
    4: "Nagoya2",
}

# engine and metadata creation
engine = create_engine(DATABASE_URL, echo=False)

# Print DB connection info for startup diagnostics (best-effort)
try:
    print(f"[DB] DATABASE_URL={DATABASE_URL}")
    print(f"[DB] engine.url={engine.url}")
    # If sqlite, show resolved filesystem path
    try:
        if DATABASE_URL.startswith("sqlite://"):
            db_file = engine.url.database
            if db_file:
                resolved = Path(db_file).expanduser().resolve()
                print(f"[DB] sqlite file={resolved}")
    except Exception:
        pass
except Exception:
    pass

# Default PDF directory for secondary PDFs when env var is not set
DEFAULT_SECONDARY_PDF_DIR = str(Path.home() / "apathy_screen_app" / "PDF")


def _ensure_table_columns(engine):
    """Ensure legacy DB has necessary columns for SecondaryArticle.
    Adds missing columns via ALTER TABLE when possible (SQLite).
    """
    required = {
        'is_physical': 'INTEGER DEFAULT 0',
        'is_brain': 'INTEGER DEFAULT 0',
        'is_psycho': 'INTEGER DEFAULT 0',
        'is_drug': 'INTEGER DEFAULT 0',
        'pdf_exists': 'INTEGER DEFAULT 0',
        'created_at': 'TEXT',
        'updated_at': 'TEXT',
    }
    try:
        with engine.connect() as conn:
            rows = conn.execute(text("PRAGMA table_info(secondaryarticle)")).all()
            existing = {r[1] for r in rows}
            for col, coldef in required.items():
                if col not in existing:
                    try:
                        conn.execute(text(f"ALTER TABLE secondaryarticle ADD COLUMN {col} {coldef}"))
                    except Exception:
                        # ignore failures (e.g., table missing) and continue
                        pass
    except Exception:
        pass


# NOTE: Do NOT run DDL (create_all / ALTER) at import time to avoid accidental
# schema changes against production DBs. Table creation / migrations should be
# executed explicitly (e.g. via migration scripts) or by setting
# AUTO_CREATE_TABLES=1 in development environments.

TEMPLATE_DIR = BASE_DIR / "templates"
app = FastAPI()

# Logger for server-side messages
logger = logging.getLogger("uvicorn.error")

templates = Jinja2Templates(directory=str(TEMPLATE_DIR))
templates.env.globals["GROUP_NAMES"] = GROUP_NAMES
# 二次スクリーニングのグループ表示ラベル（日本語）
templates.env.globals["SECONDARY_GROUP_LABELS"] = {
    "physical": "身体的要因",
    "brain": "脳神経疾患",
    "psycho": "精神・心理・環境要因",
    "drug": "薬剤性",
}

# PDF署名設定（CoreServer連携）
PDF_SECRET = os.getenv("PDF_SECRET")
CORESERVER_PDF_ENDPOINT = os.getenv("CORESERVER_PDF_ENDPOINT", "https://ifc66.v2003.coreserver.jp/pdf.php")
try:
    PDF_TTL_SEC = int(os.getenv("PDF_TTL_SEC", "300"))
except Exception:
    PDF_TTL_SEC = 300
PDF_USE_SIGNED = bool(PDF_SECRET)

if not PDF_USE_SIGNED:
    print("[PDF] PDF_SECRET not set — using local PDF endpoint /secondary/pdf/{pmid} for development")

def build_pdf_url(pmid: int) -> str:
    """Return a URL to use for PDF display/redirect.

    - When `PDF_SECRET` is configured, build a signed CoreServer URL.
    - Otherwise return the local secondary PDF endpoint which serves files from
      `SECONDARY_PDF_DIR` (useful for local development).
    """
    if PDF_USE_SIGNED:
        exp = int(time.time()) + PDF_TTL_SEC
        msg = f"{pmid}:{exp}".encode("utf-8")
        sig = hmac.new(PDF_SECRET.encode("utf-8"), msg, hashlib.sha256).hexdigest()
        return f"{CORESERVER_PDF_ENDPOINT}?pmid={pmid}&exp={exp}&sig={sig}"
    # Local fallback
    return f"/secondary/pdf/{pmid}"

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

pwd_context = CryptContext(
    schemes=["pbkdf2_sha256", "sha256_crypt"],
    deprecated="auto",
)

# =========================================================
# Middleware
# =========================================================
@app.middleware("http")
async def add_user_to_request(request: Request, call_next):
    user_id = request.session.get("user_id")
    if user_id:
        with Session(engine) as session:
            user = session.get(User, user_id)
            request.state.user = user
    else:
        request.state.user = None
    response = await call_next(request)
    return response

app.add_middleware(SessionMiddleware, secret_key="very-secret-key-for-apathy-app")

# =========================================================
# Helpers
# =========================================================
def get_current_user(request: Request) -> Optional[User]:
    return request.state.user

def get_year_min(session: Session) -> Optional[int]:
    cfg = session.get(AppConfig, 1)
    if cfg is None:
        cfg = AppConfig(id=1, year_min=2015)
        session.add(cfg)
        session.commit()
    return cfg.year_min

def set_year_min(session: Session, year_min: Optional[int]):
    cfg = session.get(AppConfig, 1)
    if cfg is None:
        cfg = AppConfig(id=1, year_min=year_min)
        session.add(cfg)
    else:
        cfg.year_min = year_min
    session.commit()

def ensure_default_users():
    with Session(engine) as session:
        if session.exec(select(func.count(User.id))).one() > 0:
            return
        users = [
            ("user1", "password1", 1, True),
            ("user2", "password2", 1, False),
            ("user3", "password3", 2, False),
            ("user4", "password4", 2, False),
            ("user5", "password5", 3, False),
            ("user6", "password6", 3, False),
            ("user7", "password7", 4, False),
            ("user8", "password8", 4, False),
        ]
        for uname, pw, grp, is_adm in users:
            u = User(username=uname, password_hash=pwd_context.hash(pw), group_no=grp, is_admin=is_adm)
            session.add(u)
        session.commit()

@app.on_event("startup")
def on_startup():
    auto = os.getenv("AUTO_CREATE_TABLES", "0")
    if auto == "1":
        # Development: create tables automatically and attempt to patch legacy schema
        print("AUTO_CREATE_TABLES=1 -> running SQLModel.metadata.create_all(engine)")
        SQLModel.metadata.create_all(engine)
        try:
            _ensure_table_columns(engine)
        except Exception:
            pass
        try:
            ensure_default_users()
        except Exception as e:
            print("Warning: ensure_default_users failed after create_all:", e)
    else:
        # Production-ish: do not modify schema. Try to ensure default users but do not fail startup.
        try:
            ensure_default_users()
        except Exception:
            print("AUTO_CREATE_TABLES!=1: skipping automatic table creation. If this is a new DB run migration script `migrate_secondary_schema`.")

    # Log whether secondaryreview table exists (read-only check). Works for SQLite.
    try:
        eng_url = str(engine.url)
        if 'sqlite' in eng_url:
            with engine.connect() as conn:
                r = conn.execute(text("SELECT name FROM sqlite_master WHERE type='table' AND name='secondaryreview' LIMIT 1")).first()
                if r:
                    print("[DB CHECK] table 'secondaryreview' exists in DB")
                else:
                    print("[DB CHECK] table 'secondaryreview' NOT found in DB")
        else:
            print("[DB CHECK] DB is not SQLite; skipping sqlite_master check")
    except Exception as e:
        print(f"[DB CHECK] failed to check secondaryreview table: {e}")

def get_group_article_ids(session: Session, year_min: Optional[int], user_group_no: int) -> List[int]:
    all_rows = list(session.exec(select(Article.id, Article.authors, Article.pmid, Article.year)))
    if not all_rows: return []
    rows = all_rows
    if year_min is not None:
        filtered = [r for r in all_rows if (r[3] is not None and r[3] >= year_min)]
        if filtered: rows = filtered
    rows.sort(key=lambda r: (r[1] or "", r[2] or 0))
    n = len(rows)
    id_list = []
    for i, row in enumerate(rows):
        if ((i * N_GROUPS) // n + 1) == user_group_no:
            id_list.append(row[0])
    return id_list


def table_has_columns(session: Session, table_name: str, col_names: List[str]) -> Dict[str, bool]:
    """
    Check whether given columns exist in a SQLite table using PRAGMA table_info.
    Returns a dict mapping column name -> bool.
    """
    try:
        rows = session.exec(text(f"PRAGMA table_info({table_name})")).all()
        existing = {r[1] for r in rows}
        return {c: (c in existing) for c in col_names}
    except Exception:
        return {c: False for c in col_names}

def get_group_scale_article_ids(session: Session, user_group_no: int) -> List[int]:
    rows = session.exec(select(ScaleArticle.id).where(ScaleArticle.group_no == user_group_no).order_by(ScaleArticle.id)).all()
    if rows: return [int(r) for r in rows]
    all_rows = list(session.exec(select(ScaleArticle.id, ScaleArticle.pmid)))
    if not all_rows: return []
    all_rows.sort(key=lambda r: (r[1] or 0))
    n = len(all_rows)
    id_list = []
    for i, row in enumerate(all_rows):
        if ((i * N_GROUPS) // n + 1) == user_group_no:
            id_list.append(row[0])
    return id_list


def get_article_safe(session: Session, article_id: int) -> Optional[SimpleNamespace]:
    """
    Fetch a minimal safe set of Article columns that likely exist in older DBs.
    Returns a SimpleNamespace with attributes for accessed fields, or None.
    """
    cols = [
        "id", "pmid", "title_en", "title_ja", "abstract_en", "abstract_ja",
        "doi", "year", "authors", "direction_gpt", "direction_gemini",
        "condition_list_gpt", "condition_list_gemini",
    ]
    existing = table_has_columns(session, "article", cols)
    fields = []
    for c in cols:
        if existing.get(c, False):
            fields.append(getattr(Article, c))

    if not fields:
        return None

    row = session.exec(select(*fields).where(Article.id == article_id)).first()
    if not row:
        return None

    data = {}
    idx = 0
    for c in cols:
        if existing.get(c, False):
            data[c] = row[idx]
            idx += 1
        else:
            data[c] = None

    return SimpleNamespace(**data)


def get_scale_article_safe(session: Session, article_id: int) -> Optional[SimpleNamespace]:
    cols = ["id", "pmid", "title_en", "title_ja", "year", "doi", "group_no"]
    existing = table_has_columns(session, "scalearticle", cols)
    fields = []
    for c in cols:
        if existing.get(c, False):
            fields.append(getattr(ScaleArticle, c))
    if not fields:
        return None
    row = session.exec(select(*fields).where(ScaleArticle.id == article_id)).first()
    if not row:
        return None
    data = {}
    idx = 0
    for c in cols:
        if existing.get(c, False):
            data[c] = row[idx]
            idx += 1
        else:
            data[c] = None
    return SimpleNamespace(**data)


def _serialize_article(article) -> Optional[dict]:
    """
    ORM Article -> template-safe dict
    """
    if not article:
        return None
    return {
        "id": getattr(article, "id", None),
        "pmid": getattr(article, "pmid", None),
        "title_en": getattr(article, "title_en", None),
        "abstract_en": getattr(article, "abstract_en", None),
        "title_ja": getattr(article, "title_ja", None),
        "abstract_ja": getattr(article, "abstract_ja", None),
        "doi": getattr(article, "doi", None),
        "year": getattr(article, "year", None),
    }


def _serialize_secondary(secondary) -> Optional[dict]:
    if not secondary:
        return None
    return {
        "pmid": getattr(secondary, "pmid", None),
        "is_physical": bool(getattr(secondary, "is_physical", False)),
        "is_brain": bool(getattr(secondary, "is_brain", False)),
        "is_psycho": bool(getattr(secondary, "is_psycho", False)),
        "is_drug": bool(getattr(secondary, "is_drug", False)),
        "pdf_exists": bool(getattr(secondary, "pdf_exists", False)),
    }


def _serialize_auto(auto) -> Optional[dict]:
    if not auto:
        return None
    return {
        "pmid": getattr(auto, "pmid", None),
        "auto_target_condition": getattr(auto, "auto_target_condition", None),
        "auto_apathy_terms": getattr(auto, "auto_apathy_terms", None),
        "auto_population_N": getattr(auto, "auto_population_N", None),
        "auto_prevalence": getattr(auto, "auto_prevalence", None),
        "auto_intervention": getattr(auto, "auto_intervention", None),
        "auto_confidence": getattr(auto, "auto_confidence", None),
        "needs_review": getattr(auto, "needs_review", False),
    }


def _serialize_review(review) -> Optional[dict]:
    if not review:
        return None
    return {
        "pmid": getattr(review, "pmid", None),
        "group": getattr(review, "group", None),
        "reviewer_id": getattr(review, "reviewer_id", None),
        "decision": getattr(review, "decision", None),
        "final_apathy_terms": getattr(review, "final_apathy_terms", None),
        "final_target_condition": getattr(review, "final_target_condition", None),
        "final_population_n": getattr(review, "final_population_n", None),
        "final_prevalence": getattr(review, "final_prevalence", None),
        "final_intervention": getattr(review, "final_intervention", None),
        "comment": getattr(review, "comment", None),
    }

def check_group_status(session: Session, group_no: int, mode: str = "disease"):
    """
    指定グループの進捗状態とコンフリクト有無をチェックする
    """
    users = session.exec(select(User).where(User.group_no == group_no)).all()
    if not users: return False, False

    if mode == "disease":
        year_min = get_year_min(session)
        article_ids = get_group_article_ids(session, year_min, group_no)
        # Select only minimal columns to avoid querying possibly-missing columns
        decisions = session.exec(select(ScreeningDecision.article_id, ScreeningDecision.user_id, ScreeningDecision.decision).where(ScreeningDecision.article_id.in_(article_ids))).all()
        decision_map = defaultdict(dict)
        for aid, uid, dec in decisions:
            if dec is not None:
                decision_map[aid][uid] = dec
    else: # scale
        article_ids = get_group_scale_article_ids(session, group_no)
        decisions = session.exec(select(ScaleScreeningDecision.scale_article_id, ScaleScreeningDecision.user_id, ScaleScreeningDecision.rating).where(ScaleScreeningDecision.scale_article_id.in_(article_ids))).all()
        decision_map = defaultdict(dict)
        for aid, uid, rating in decisions:
            if rating is not None:
                decision_map[aid][uid] = rating

    total_articles = len(article_ids)
    if total_articles == 0: return False, False

    # 全員完了チェック
    for u in users:
        user_done_count = 0
        for aid in article_ids:
            if u.id in decision_map.get(aid, {}):
                user_done_count += 1
        if user_done_count < total_articles:
            return False, False

    # コンフリクトチェック
    has_conflicts = False
    for aid in article_ids:
        votes = decision_map.get(aid, {})
        vals = list(votes.values())
        
        has_2 = any(v == 2 for v in vals)
        has_1 = any(v == 1 for v in vals)
        has_0 = any(v == 0 for v in vals)
        
        if not has_2:
            if has_1 and has_0:
                has_conflicts = True
                break
    
    return True, has_conflicts

# =========================================================
# Routes: Common
# =========================================================
@app.get("/", response_class=HTMLResponse, name="index")
def index(request: Request):
    user = get_current_user(request)
    return templates.TemplateResponse("index.html", {
        "request": request, "username": user.username if user else None,
        "group_no": user.group_no if user else None, "current_page": "home"
    })

@app.get("/database", response_class=HTMLResponse)
def database(request: Request):
    user = get_current_user(request)
    with Session(engine) as session:
        # select minimal columns to avoid errors when newer Article columns are missing
        rows = session.exec(select(Article.id, Article.pmid, Article.title_en, Article.title_ja, Article.year).order_by(Article.id)).all()
        articles = [SimpleNamespace(id=r[0], pmid=r[1], title_en=r[2], title_ja=r[3], year=r[4]) for r in rows]
    return templates.TemplateResponse("database.html", {
        "request": request, "articles": articles,
        "username": user.username if user else None,
        "group_no": user.group_no if user else None, "current_page": "database"
    })

@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    is_admin = user.is_admin
    
    current_step = 1
    year_min = 2015

    with Session(engine) as session:
        year_min = get_year_min(session)
        # 進捗チェックのみ実施 (Step 1 -> 2)
        users = session.exec(select(User)).all()
        total_assigned = 0
        total_done = 0
        for u in users:
            ids = get_group_article_ids(session, year_min, u.group_no)
            done = session.exec(select(func.count(ScreeningDecision.id)).where(
                (ScreeningDecision.user_id == u.id) & 
                (ScreeningDecision.article_id.in_(ids)) & 
                (ScreeningDecision.decision.is_not(None))
            )).one()
            total_assigned += len(ids)
            total_done += done
        if total_assigned > 0 and (total_done / total_assigned) > 0.98:
            current_step = 2

    return templates.TemplateResponse("settings.html", {
        "request": request, "username": user.username, "group_no": user.group_no,
        "year_min": year_min, "is_admin": is_admin, 
        "current_step": current_step,
        "current_page": "settings"
    })

@app.post("/settings", response_class=HTMLResponse)
def settings_submit(request: Request, year_min: Optional[int] = Form(None)):
    user = get_current_user(request)
    if not user or not user.is_admin: return RedirectResponse("/settings", 303)
    with Session(engine) as session:
        set_year_min(session, year_min)
    return RedirectResponse("/settings", 303)

@app.get("/admin/users", response_class=HTMLResponse)
def admin_users_page(request: Request):
    user = get_current_user(request)
    if not user or not user.is_admin: return RedirectResponse("/settings", 303)
    with Session(engine) as session:
        all_users = session.exec(select(User).order_by(User.id)).all()
    return templates.TemplateResponse("admin_users.html", {
        "request": request, "username": user.username, "group_no": user.group_no,
        "all_users": all_users, "current_page": "settings"
    })

@app.post("/admin/users/update", response_class=HTMLResponse)
def admin_user_update(request: Request, user_id: int = Form(...), username: str = Form(...), group_no: int = Form(...), is_admin: bool = Form(False)):
    current_user = get_current_user(request)
    if not current_user or not current_user.is_admin: return RedirectResponse("/settings", 303)
    with Session(engine) as session:
        target_user = session.get(User, user_id)
        if target_user:
            target_user.username = username
            target_user.group_no = group_no
            if target_user.id != current_user.id:
                target_user.is_admin = is_admin
            else:
                target_user.is_admin = True
            session.add(target_user)
            session.commit()
    return RedirectResponse("/admin/users", 303)

# =========================================================
# Routes: Auth
# =========================================================
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, next: str = Query("screen")):
    return templates.TemplateResponse("login.html", {"request": request, "error": None, "next": next, "current_page": "login"})

@app.post("/login", response_class=HTMLResponse)
def login(request: Request, username: str = Form(...), password: str = Form(...), next: str = Form("screen")):
    with Session(engine) as session:
        user = session.exec(select(User).where(User.username == username)).first()
        if not user or not pwd_context.verify(password, user.password_hash):
            return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid credentials", "next": next})
        request.session["user_id"] = user.id
    if next == "scale": return RedirectResponse("/scale_screen", 303)
    if next == "conflicts": return RedirectResponse("/conflicts", 303)
    return RedirectResponse("/screen", 303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/", 303)

@app.get("/change_password", response_class=HTMLResponse)
def change_password_page(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    return templates.TemplateResponse("change_password.html", {"request": request, "error": None, "success": None, "username": user.username, "group_no": user.group_no, "current_page": "change_password"})

@app.post("/change_password", response_class=HTMLResponse)
def change_password(request: Request, current_password: str = Form(...), new_password: str = Form(...), new_password_confirm: str = Form(...)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    if new_password != new_password_confirm:
        return templates.TemplateResponse("change_password.html", {"request": request, "error": "Passwords do not match", "success": None, "username": user.username, "group_no": user.group_no, "current_page": "change_password"})
    if len(new_password) < 6:
        return templates.TemplateResponse("change_password.html", {"request": request, "error": "Password too short", "success": None, "username": user.username, "group_no": user.group_no, "current_page": "change_password"})
    with Session(engine) as session:
        db_user = session.exec(select(User).where(User.id == user.id)).first()
        if not db_user or not pwd_context.verify(current_password, db_user.password_hash):
            return templates.TemplateResponse("change_password.html", {"request": request, "error": "Incorrect current password", "success": None, "username": user.username, "group_no": user.group_no, "current_page": "change_password"})
        db_user.password_hash = pwd_context.hash(new_password)
        session.add(db_user)
        session.commit()
    return templates.TemplateResponse("change_password.html", {"request": request, "error": None, "success": "Password changed", "username": user.username, "group_no": user.group_no, "current_page": "change_password"})

# =========================================================
# Routes: Disease Screen
# =========================================================
@app.get("/screen", response_class=HTMLResponse, name="screen_page")
def screen_page(request: Request, group_no: int = Query(None), article_index: int = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    user_id = user.id
    group_no = user.group_no if group_no is None else group_no

    with Session(engine) as session:
        year_min = get_year_min(session)
        id_list = get_group_article_ids(session, year_min, group_no)
        total = len(id_list)
        rated = session.exec(select(func.count(ScreeningDecision.id)).where(
            (ScreeningDecision.user_id == user_id) & (ScreeningDecision.article_id.in_(id_list)) & (ScreeningDecision.decision.is_not(None))
        )).one() if id_list else 0

        article = None
        current_index = None
        if id_list:
            if article_index is not None:
                idx = max(1, min(article_index, total))
                article = get_article_safe(session, id_list[idx - 1])
                current_index = idx
            else:
                decided_ids = set(session.exec(select(ScreeningDecision.article_id).where(ScreeningDecision.user_id == user_id)).all())
                for i, aid in enumerate(id_list):
                    if aid not in decided_ids:
                        article = get_article_safe(session, aid)
                        current_index = i + 1
                        break
                if not article:
                    article = get_article_safe(session, id_list[-1])
                    current_index = total

        prev_decision = None
        prev_comment = ""
        prev_flag_cause = False
        prev_flag_treatment = False
        prev_cat_physical = False
        prev_cat_brain = False
        prev_cat_psycho = False
        prev_cat_drug = False

        if article:
            # select only columns that exist to avoid missing-column errors
            cat_cols = ["cat_physical", "cat_brain", "cat_psycho", "cat_drug"]
            has_cols = table_has_columns(session, "screeningdecision", cat_cols)
            fields = [ScreeningDecision.decision, ScreeningDecision.comment, ScreeningDecision.flag_cause, ScreeningDecision.flag_treatment]
            for c in cat_cols:
                if has_cols.get(c, False):
                    fields.append(getattr(ScreeningDecision, c))

            row = session.exec(select(*fields).where((ScreeningDecision.user_id == user_id) & (ScreeningDecision.article_id == article.id))).first()
            if row:
                prev_decision = row[0]
                prev_comment = row[1] or ""
                prev_flag_cause = bool(row[2])
                prev_flag_treatment = bool(row[3])
                offset = 4
                prev_cat_physical = bool(row[offset]) if has_cols.get("cat_physical", False) else False
                prev_cat_brain = bool(row[offset + (1 if has_cols.get("cat_physical", False) else 0)]) if has_cols.get("cat_brain", False) else False
                # compute offsets more robustly
                idx = 4
                if has_cols.get("cat_physical", False):
                    prev_cat_physical = bool(row[idx]); idx += 1
                if has_cols.get("cat_brain", False):
                    prev_cat_brain = bool(row[idx]); idx += 1
                if has_cols.get("cat_psycho", False):
                    prev_cat_psycho = bool(row[idx]); idx += 1
                if has_cols.get("cat_drug", False):
                    prev_cat_drug = bool(row[idx]); idx += 1

    return templates.TemplateResponse("screen.html", {
        "request": request, "group_no": group_no, "username": user.username,
        "article": article, "direction_memo": None, "progress_done": rated, "progress_total": total,
        "current_index": current_index, "current_page": "screen",
        "prev_decision": prev_decision, "prev_comment": prev_comment,
        "prev_flag_cause": prev_flag_cause, "prev_flag_treatment": prev_flag_treatment,
        "prev_cat_physical": prev_cat_physical,
        "prev_cat_brain": prev_cat_brain,
        "prev_cat_psycho": prev_cat_psycho,
        "prev_cat_drug": prev_cat_drug,
    })

@app.post("/screen", response_class=HTMLResponse, name="submit_screen")
def submit_screen(
    request: Request, article_id: int = Form(...), decision: Optional[int] = Form(None),
    flag_cause: int = Form(0), flag_treatment: int = Form(0),
    cat_physical: int = Form(0), cat_brain: int = Form(0), cat_psycho: int = Form(0),
    cat_drug: int = Form(0),
    nav: str = Form("next"),
    jump_index: str | None = Form(None), comment: str = Form("")
):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    
    with Session(engine) as session:
        target_article = get_article_safe(session, article_id)
        target_group_no = target_article.group_no if target_article and getattr(target_article, 'group_no', None) is not None else (user.group_no or 1)
        year_min = get_year_min(session)
        id_list = get_group_article_ids(session, year_min, target_group_no)
        try: current_index = id_list.index(article_id) + 1
        except ValueError: current_index = 1
        
        # Use core UPDATE/INSERT limiting to columns that actually exist in DB
        cat_cols = ["cat_physical", "cat_brain", "cat_psycho", "cat_drug"]
        has_cols = table_has_columns(session, "screeningdecision", cat_cols)

        existing_id = session.exec(select(ScreeningDecision.id).where((ScreeningDecision.user_id == user.id) & (ScreeningDecision.article_id == article_id))).first()

        if existing_id:
            # build UPDATE with only existing columns
            set_clauses = []
            params = {"id": int(existing_id)}
            if decision is not None:
                set_clauses.append("decision = :decision")
                params["decision"] = int(decision)
            set_clauses.append("comment = :comment")
            params["comment"] = comment or ""
            set_clauses.append("flag_cause = :flag_cause")
            params["flag_cause"] = int(bool(flag_cause))
            set_clauses.append("flag_treatment = :flag_treatment")
            params["flag_treatment"] = int(bool(flag_treatment))
            # category flags only when present
            if has_cols.get("cat_physical", False):
                set_clauses.append("cat_physical = :cat_physical")
                params["cat_physical"] = int(bool(cat_physical))
            if has_cols.get("cat_brain", False):
                set_clauses.append("cat_brain = :cat_brain")
                params["cat_brain"] = int(bool(cat_brain))
            if has_cols.get("cat_psycho", False):
                set_clauses.append("cat_psycho = :cat_psycho")
                params["cat_psycho"] = int(bool(cat_psycho))
            if has_cols.get("cat_drug", False):
                set_clauses.append("cat_drug = :cat_drug")
                params["cat_drug"] = int(bool(cat_drug))

            if set_clauses:
                sql = f"UPDATE screeningdecision SET {', '.join(set_clauses)} WHERE id = :id"
                session.exec(text(sql), params)
                session.commit()
        else:
            # INSERT only when user provided a decision
            if decision is not None:
                cols = ["user_id", "article_id", "decision", "comment", "flag_cause", "flag_treatment"]
                vals = [":user_id", ":article_id", ":decision", ":comment", ":flag_cause", ":flag_treatment"]
                params = {
                    "user_id": user.id,
                    "article_id": article_id,
                    "decision": int(decision),
                    "comment": comment or "",
                    "flag_cause": int(bool(flag_cause)),
                    "flag_treatment": int(bool(flag_treatment)),
                }
                if has_cols.get("cat_physical", False):
                    cols.append("cat_physical"); vals.append(":cat_physical"); params["cat_physical"] = int(bool(cat_physical))
                if has_cols.get("cat_brain", False):
                    cols.append("cat_brain"); vals.append(":cat_brain"); params["cat_brain"] = int(bool(cat_brain))
                if has_cols.get("cat_psycho", False):
                    cols.append("cat_psycho"); vals.append(":cat_psycho"); params["cat_psycho"] = int(bool(cat_psycho))
                if has_cols.get("cat_drug", False):
                    cols.append("cat_drug"); vals.append(":cat_drug"); params["cat_drug"] = int(bool(cat_drug))

                sql = f"INSERT INTO screeningdecision ({', '.join(cols)}) VALUES ({', '.join(vals)})"
                session.exec(text(sql), params)
                session.commit()
    
    total = len(id_list)
    target = current_index
    if nav == "prev": target = max(1, current_index - 1)
    elif nav == "jump" and jump_index: 
        try: target = max(1, min(int(jump_index), total))
        except: pass
    elif nav == "next" and total > 0: target = min(current_index + 1, total)
    
    return RedirectResponse(f"/screen?article_index={target}&group_no={target_group_no}", 303)

# =========================================================
# Routes: Scale Screen
# =========================================================
@app.get("/scale_screen", response_class=HTMLResponse, name="scale_screen_page")
def scale_screen_page(request: Request, article_index: int = Query(1, ge=1), group_no: Optional[int] = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login?next=scale", 303)
    user_id = user.id
    group_no = user.group_no if group_no is None else group_no

    with Session(engine) as session:
        id_list = get_group_scale_article_ids(session, group_no)
        total = len(id_list)
        my_done = session.exec(select(func.count(ScaleScreeningDecision.id)).where(
            (ScaleScreeningDecision.user_id == user_id) & (ScaleScreeningDecision.scale_article_id.in_(id_list)) & (ScaleScreeningDecision.rating.is_not(None))
        )).one() if id_list else 0
        
        total_all = session.exec(select(func.count(ScaleArticle.id))).one()
        all_done = session.exec(select(func.count(func.distinct(ScaleScreeningDecision.scale_article_id))).where(ScaleScreeningDecision.rating.is_not(None))).one()

        article = None
        current_index = None
        if id_list:
            if article_index is not None:
                idx = max(1, min(article_index, total))
                article = session.get(ScaleArticle, id_list[idx - 1])
                current_index = idx
            else:
                decided = set(session.exec(select(ScaleScreeningDecision.scale_article_id).where(ScaleScreeningDecision.user_id == user_id)).all())
                for i, aid in enumerate(id_list):
                    if aid not in decided:
                        article = session.get(ScaleArticle, aid)
                        current_index = i + 1
                        break
                if not article:
                    article = session.get(ScaleArticle, id_list[-1])
                    current_index = total

        if article:
            if not article.doi and article.pmid:
                found_doi = session.exec(select(Article.doi).where(Article.pmid == article.pmid)).first()
                if found_doi:
                    article.doi = found_doi

        my_rating = None
        my_comment = ""
        if article:
            existing = session.exec(select(ScaleScreeningDecision).where((ScaleScreeningDecision.user_id == user_id) & (ScaleScreeningDecision.scale_article_id == article.id))).first()
            if existing:
                my_rating = existing.rating
                my_comment = existing.comment or ""

    return templates.TemplateResponse("scale_screen.html", {
        "request": request, "username": user.username, "group_no": group_no,
        "article": article, "progress_done": my_done, "progress_total": total,
        "current_index": current_index, "scale_all_done": all_done, "scale_total_all": total_all,
        "my_rating": my_rating, "my_comment": my_comment, "current_page": "scale_screen"
    })

@app.post("/scale_screen", response_class=HTMLResponse, name="submit_scale_screen")
def submit_scale_screen(
    request: Request, article_id: int = Form(...), decision: Optional[int] = Form(None),
    comment: str = Form(""), nav: str = Form("next"), jump_index: Optional[str] = Form(None)
):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login?next=scale", 303)
    
    with Session(engine) as session:
        target_article = session.get(ScaleArticle, article_id)
        target_group_no = target_article.group_no if target_article else user.group_no

        id_list = get_group_scale_article_ids(session, target_group_no)
        try: current_index = id_list.index(article_id) + 1
        except ValueError: current_index = 1
        
        existing = session.exec(select(ScaleScreeningDecision).where((ScaleScreeningDecision.user_id == user.id) & (ScaleScreeningDecision.scale_article_id == article_id))).first()
        if existing:
            if decision is not None: existing.rating = decision
            if comment is not None: existing.comment = comment
            session.add(existing)
        elif decision is not None:
            session.add(ScaleScreeningDecision(user_id=user.id, scale_article_id=article_id, rating=decision, comment=comment or ""))
        session.commit()

    target = current_index
    total = len(id_list)
    if nav == "prev": target = max(1, current_index - 1)
    elif nav == "jump" and jump_index:
        try: target = max(1, min(int(jump_index), total))
        except: pass
    elif nav == "next" and total > 0: target = min(current_index + 1, total)

    return RedirectResponse(f"/scale_screen?article_index={target}&group_no={target_group_no}", 303)

# =========================================================
# Routes: Progress Pages
# =========================================================
@app.get("/my_index", response_class=HTMLResponse, name="disease_progress_my")
def my_index(request: Request, target_user_id: Optional[int] = Query(None)):
    current_user = get_current_user(request)
    if not current_user: return RedirectResponse("/login", 303)
    view_user = current_user
    if target_user_id:
        with Session(engine) as session:
            u = session.get(User, target_user_id)
            if u: view_user = u
    request.state.user = current_user
    with Session(engine) as session:
        year_min = get_year_min(session)
        id_list = get_group_article_ids(session, year_min, view_user.group_no)
        done_count = session.exec(select(func.count(ScreeningDecision.id)).where(
            (ScreeningDecision.user_id == view_user.id) & (ScreeningDecision.article_id.in_(id_list)) & (ScreeningDecision.decision.is_not(None))
        )).one() if id_list else 0
        # select minimal Article columns + decision to avoid loading missing Article columns
        fields = [Article.id, Article.pmid, Article.title_en, Article.title_ja, ScreeningDecision.decision]
        stmt = select(*fields).join(
            ScreeningDecision,
            (ScreeningDecision.article_id == Article.id) & (ScreeningDecision.user_id == view_user.id),
            isouter=True,
        ).where(Article.id.in_(id_list)).order_by(Article.id)
        raw = session.exec(stmt).all()
        rows = []
        for r in raw:
            art = SimpleNamespace(id=r[0], pmid=r[1], title_en=r[2], title_ja=r[3])
            dec = SimpleNamespace(decision=r[4]) if r[4] is not None else None
            rows.append((art, dec))
    return templates.TemplateResponse("my_index.html", {
        "request": request, "rows": rows, "username": current_user.username, "group_no": current_user.group_no,
        "view_user": view_user, "current_page": "my_index", 
        "progress_done": done_count, "progress_total": len(id_list), "progress_pct": int(done_count * 100 / len(id_list)) if id_list else 0
    })

@app.get("/scale_my_index", response_class=HTMLResponse, name="scale_progress_my")
def scale_my_index(request: Request, target_user_id: Optional[int] = Query(None)):
    current_user = get_current_user(request)
    if not current_user: return RedirectResponse("/login?next=scale", 303)
    view_user = current_user
    if target_user_id:
        with Session(engine) as session:
            u = session.get(User, target_user_id)
            if u: view_user = u
    with Session(engine) as session:
        id_list = get_group_scale_article_ids(session, view_user.group_no)
        done_count = session.exec(select(func.count(ScaleScreeningDecision.id)).where(
            (ScaleScreeningDecision.user_id == view_user.id) & (ScaleScreeningDecision.scale_article_id.in_(id_list)) & (ScaleScreeningDecision.rating.is_not(None))
        )).one() if id_list else 0
        fields = [ScaleArticle.id, ScaleArticle.pmid, ScaleArticle.title_en, ScaleArticle.title_ja, ScaleScreeningDecision.rating]
        stmt = select(*fields).join(
            ScaleScreeningDecision,
            (ScaleScreeningDecision.scale_article_id == ScaleArticle.id) & (ScaleScreeningDecision.user_id == view_user.id),
            isouter=True,
        ).where(ScaleArticle.id.in_(id_list)).order_by(ScaleArticle.id)
        raw = session.exec(stmt).all()
        rows = []
        for r in raw:
            art = SimpleNamespace(id=r[0], pmid=r[1], title_en=r[2], title_ja=r[3])
            dec = SimpleNamespace(rating=r[4]) if r[4] is not None else None
            rows.append((art, dec))
    return templates.TemplateResponse("scale_my_index.html", {
        "request": request, "rows": rows, "username": current_user.username, "group_no": current_user.group_no,
        "view_user": view_user, "current_page": "scale_my_index", 
        "progress_done": done_count, "progress_total": len(id_list), 
        "progress_pct": int(done_count * 100 / len(id_list)) if id_list else 0
    })

@app.get("/dashboard", response_class=HTMLResponse, name="dashboard")
def dashboard(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    with Session(engine) as session:
        year_min = get_year_min(session)
        users = session.exec(select(User)).all()
        rows, o_d_t, o_d_r, o_s_t, o_s_r = [], 0, 0, 0, 0
        group_progress = defaultdict(lambda: {"total": 0, "done": 0})
        for u in users:
            d_ids = get_group_article_ids(session, year_min, u.group_no)
            d_r = session.exec(select(func.count(ScreeningDecision.id)).where((ScreeningDecision.user_id == u.id) & (ScreeningDecision.article_id.in_(d_ids)) & (ScreeningDecision.decision.is_not(None)))).one() if d_ids else 0
            s_ids = session.exec(select(ScaleArticle.id).where(ScaleArticle.group_no == u.group_no)).all()
            s_r = session.exec(select(func.count(ScaleScreeningDecision.id)).where((ScaleScreeningDecision.user_id == u.id) & (ScaleScreeningDecision.scale_article_id.in_(s_ids)) & (ScaleScreeningDecision.rating.is_not(None)))).one() if s_ids else 0
            
            rows.append({"id": u.id, "username": u.username, "group_no": u.group_no, "dis_rated": d_r, "dis_total": len(d_ids), "dis_pct": (d_r/len(d_ids)*100) if d_ids else 0, "scale_rated": s_r, "scale_total": len(s_ids), "scale_pct": (s_r/len(s_ids)*100) if s_ids else 0})
            o_d_t += len(d_ids); o_d_r += d_r; o_s_t += len(s_ids); o_s_r += s_r
            
            group_progress[u.group_no]["total"] += len(d_ids)
            group_progress[u.group_no]["done"] += d_r
        
        # 自分のグループの完了状態チェック
        is_disease_complete, has_disease_conflicts = check_group_status(session, user.group_no, "disease")
        is_scale_complete, has_scale_conflicts = check_group_status(session, user.group_no, "scale")

    return templates.TemplateResponse("dashboard.html", {
        "request": request, "rows": rows, "username": user.username, "group_no": user.group_no, "current_page": "dashboard",
        "overall_dis_total": o_d_t, "overall_dis_rated": o_d_r, "overall_dis_pct": (o_d_r/o_d_t*100) if o_d_t else 0,
        "overall_scale_total": o_s_t, "overall_scale_rated": o_s_r, "overall_scale_pct": (o_s_r/o_s_t*100) if o_s_t else 0,
        
        "is_disease_complete": is_disease_complete, "has_disease_conflicts": has_disease_conflicts,
        "is_scale_complete": is_scale_complete, "has_scale_conflicts": has_scale_conflicts
    })

# =========================================================
# Routes: Conflicts Resolution
# =========================================================
@app.get("/conflicts", response_class=HTMLResponse, name="conflicts_page")
def conflicts_page(request: Request, mode: str = Query("disease"), group_no: Optional[int] = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login?next=conflicts", 303)
    if group_no is None: group_no = user.group_no
    
    conflicts_list = []
    
    with Session(engine) as session:
        is_complete, has_conflicts = check_group_status(session, group_no, mode)
        
        if is_complete:
            if mode == "disease":
                year_min = get_year_min(session)
                article_ids = get_group_article_ids(session, year_min, group_no)
                # select minimal columns to avoid referencing possibly-missing cat_* columns
                decisions = session.exec(
                    select(
                        ScreeningDecision.article_id,
                        ScreeningDecision.user_id,
                        ScreeningDecision.decision,
                        ScreeningDecision.comment,
                    ).where(ScreeningDecision.article_id.in_(article_ids))
                ).all()
                art_map = defaultdict(dict)
                for aid, uid, dec, comment in decisions:
                    if dec is not None:
                        u = session.get(User, uid)
                        uname = u.username if u else str(uid)
                        art_map[aid][uname] = dec
                        art_map[aid]['_comment_' + uname] = comment
                for aid, votes in art_map.items():
                    vals = [v for k, v in votes.items() if not k.startswith('_')]
                    has_2 = any(v == 2 for v in vals)
                    has_1 = any(v == 1 for v in vals)
                    has_0 = any(v == 0 for v in vals)
                    if not has_2:
                        if has_1 and has_0:
                            art = get_article_safe(session, aid)
                            if art: conflicts_list.append({
                                "id": art.id,
                                "pmid": art.pmid,
                                "doi": art.doi,
                                "title": (art.title_ja or art.title_en) if art else None,
                                "abstract": (art.abstract_ja or art.abstract_en) if art else None,
                                "votes": votes
                            })
            else:
                scale_ids = get_group_scale_article_ids(session, group_no)
                decisions = session.exec(
                    select(
                        ScaleScreeningDecision.scale_article_id,
                        ScaleScreeningDecision.user_id,
                        ScaleScreeningDecision.rating,
                        ScaleScreeningDecision.comment,
                    ).where(ScaleScreeningDecision.scale_article_id.in_(scale_ids))
                ).all()
                art_map = defaultdict(dict)
                for aid, uid, rating, comment in decisions:
                    if rating is not None:
                        u = session.get(User, uid)
                        uname = u.username if u else str(uid)
                        art_map[aid][uname] = rating
                        art_map[aid]['_comment_' + uname] = comment
                for aid, votes in art_map.items():
                    vals = [v for k, v in votes.items() if not k.startswith('_')]
                    has_2 = any(v == 2 for v in vals)
                    has_1 = any(v == 1 for v in vals)
                    has_0 = any(v == 0 for v in vals)
                    if not has_2:
                        if has_1 and has_0:
                            art = session.get(ScaleArticle, aid)
                            if art: conflicts_list.append({
                                "id": art.id,
                                "pmid": art.pmid,
                                "doi": art.doi,
                                "title": art.title_ja or art.title_en,
                                "abstract": art.abstract_ja or art.abstract_en,
                                "votes": votes
                            })

    return templates.TemplateResponse("conflicts.html", {
        "request": request, "username": user.username, "group_no": user.group_no,
        "mode": mode, "target_group_no": group_no, "conflicts": conflicts_list,
        "is_complete": is_complete,
        "current_page": "conflicts"
    })

@app.post("/resolve_conflict", response_class=HTMLResponse)
def resolve_conflict(request: Request, mode: str = Form(...), article_id: int = Form(...), resolution: int = Form(...), target_group_no: int = Form(...)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    with Session(engine) as session:
        if mode == "disease":
            decisions = session.exec(select(ScreeningDecision).where(ScreeningDecision.article_id == article_id)).all()
            for d in decisions:
                d.decision = resolution
                session.add(d)
        else:
            decisions = session.exec(select(ScaleScreeningDecision).where(ScaleScreeningDecision.scale_article_id == article_id)).all()
            for d in decisions:
                d.rating = resolution
                session.add(d)
        session.commit()
    return RedirectResponse(f"/conflicts?mode={mode}&group_no={target_group_no}", 303)

@app.get("/export_secondary_candidates", response_class=StreamingResponse)
def export_secondary_candidates_txt(request: Request, mode: str = Query("disease"), group_no: Optional[int] = Query(None)):
    """
    Export secondary screening candidate PMIDs.
    - If `group_no` is omitted (None), export across all groups (all articles meeting year_min).
    - If `group_no` provided, behave exactly as before (group-sliced selection).
    - Prefer `Article.final_decision` when the column exists in the DB; otherwise fall back to aggregated per-user votes.
    """
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    output = io.StringIO()
    with Session(engine) as session:
        # If a specific group is requested, require that group's screening is complete and conflict-free.
        # When exporting across all groups (group_no is None), proceed regardless of completion so
        # users can download the union of current candidates.
        if group_no is not None:
            is_complete, has_conflicts = check_group_status(session, group_no, mode)
            if not is_complete or has_conflicts:
                return HTMLResponse("Error: Screening incomplete or conflicts exist.", status_code=400)

        pmids = set()

        if mode == "disease":
            year_min = get_year_min(session)

            # determine article_ids (group-sliced or all)
            if group_no is None:
                all_rows = list(session.exec(select(Article.id, Article.year)))
                if year_min is not None:
                    rows = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
                    if not rows:
                        rows = all_rows
                else:
                    rows = all_rows
                article_ids = [r[0] for r in rows]
            else:
                article_ids = get_group_article_ids(session, year_min, group_no)

            # If DB has final_decision column, use it (safe SELECT only when column exists)
            cols = table_has_columns(session, "article", ["final_decision"]) if article_ids else {"final_decision": False}
            use_final = cols.get("final_decision", False)

            if use_final:
                rows = session.exec(select(Article.id, Article.pmid, Article.final_decision).where(Article.id.in_(article_ids))).all()
                for r in rows:
                    aid, pmid, final_dec = r
                    try:
                        if final_dec is not None and int(final_dec) >= 1 and pmid:
                            pmids.add(pmid)
                    except Exception:
                        continue
            else:
                # fall back to aggregated per-user votes
                decisions = session.exec(select(ScreeningDecision.article_id, ScreeningDecision.decision).where(ScreeningDecision.article_id.in_(article_ids))).all()
                art_map = defaultdict(list)
                for aid, dec in decisions:
                    if dec is not None:
                        art_map[aid].append(int(dec))

                for aid in article_ids:
                    decs = art_map.get(aid, [])
                    if decs and max(decs) >= 1:
                        # safe to SELECT only id,pmid
                        row = session.exec(select(Article.id, Article.pmid).where(Article.id == aid)).first()
                        if row:
                            _, pmid = row
                            if pmid:
                                pmids.add(pmid)
        else:
            # scale: behave as before, group_no None => all scale articles
            if group_no is None:
                scale_rows = list(session.exec(select(ScaleArticle.id)))
                scale_ids = [r[0] for r in scale_rows]
            else:
                scale_ids = get_group_scale_article_ids(session, group_no)

            decisions = session.exec(select(ScaleScreeningDecision.scale_article_id, ScaleScreeningDecision.rating).where(ScaleScreeningDecision.scale_article_id.in_(scale_ids))).all()
            art_map = defaultdict(list)
            for aid, rating in decisions:
                if rating is not None:
                    art_map[aid].append(int(rating))

            for aid in scale_ids:
                decs = art_map.get(aid, [])
                if decs and max(decs) >= 1:
                    row = session.exec(select(ScaleArticle.id, ScaleArticle.pmid).where(ScaleArticle.id == aid)).first()
                    if row:
                        _, pmid = row
                        if pmid:
                            pmids.add(pmid)

        for pmid in sorted(list(pmids)):
            output.write(f"{pmid}\n")
    output.seek(0)

    # filename: include allgroups marker when group_no omitted
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    if group_no is None:
        filename = f"secondary_candidates_{mode}_allgroups_{ts}.txt"
    else:
        filename = f"secondary_candidates_{mode}_g{group_no}_{ts}.txt"

    return StreamingResponse(output, media_type="text/plain; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/export_secondary_pmid_list", response_class=StreamingResponse)
def export_secondary_pmid_list(request: Request, mode: str = Query("disease"), group_no: Optional[int] = Query(None)):
    """
    Export CSV list of PMIDs for secondary screening based on aggregated ScreeningDecision.
    - Does NOT alter DB schema.
    - If `group_no` omitted -> all groups (respecting `year_min` filter).
    - Aggregates `ScreeningDecision` by `article_id` and picks the mode (most frequent) decision.
      Ties => `PENDING`.
    - Outputs only articles whose final aggregated decision is not an exclusion (safe default: decision != 0).
    """
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", 303)

    output = io.StringIO()
    writer = csv.writer(output)
    # header
    writer.writerow(["pmid", "decision_final", "group_no", "year", "title_en", "title_ja"])

    with Session(engine) as session:
        # determine article ids to consider
        year_min = get_year_min(session)
        if mode != "disease":
            # For now only disease supported
            return HTMLResponse("Only disease mode is supported for this export.", status_code=400)

        if group_no is None:
            all_rows = list(session.exec(select(Article.id, Article.year)))
            if year_min is not None:
                rows = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
                if not rows:
                    rows = all_rows
            else:
                rows = all_rows
            article_ids = [r[0] for r in rows]
        else:
            article_ids = get_group_article_ids(session, year_min, group_no)

        if not article_ids:
            output.seek(0)
            ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            if group_no is None:
                filename = f"secondary_pmid_list_{mode}_allgroups_{ts}.csv"
            else:
                filename = f"secondary_pmid_list_{mode}_g{group_no}_{ts}.csv"
            return StreamingResponse(output, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

        # Aggregate screeningdecision counts per article_id,decision
        sd_rows = session.exec(
            select(ScreeningDecision.article_id, ScreeningDecision.decision, func.count())
            .where(ScreeningDecision.article_id.in_(article_ids))
            .group_by(ScreeningDecision.article_id, ScreeningDecision.decision)
        ).all()

        from collections import defaultdict

        counts = defaultdict(lambda: defaultdict(int))
        for aid, dec, c in sd_rows:
            if dec is None:
                continue
            counts[aid][str(int(dec))] += int(c)

        # decide final per-article
        final_map = {}
        for aid in article_ids:
            decs = counts.get(aid, {})
            if not decs:
                final_map[aid] = "PENDING"
                continue
            maxc = max(decs.values())
            winners = [d for d, cnt in decs.items() if cnt == maxc]
            if len(winners) != 1:
                final_map[aid] = "PENDING"
            else:
                final_map[aid] = winners[0]

        # treat '0' as exclusion by default; everything else goes to secondary candidates
        EXCLUDE_DECISIONS = {"0"}
        DECISION_LABEL = {"0": "exclude", "1": "include", "2": "hold", "PENDING": "PENDING"}

        # Fetch minimal article fields for those selected
        for aid, fin in final_map.items():
            label = DECISION_LABEL.get(fin, fin)
            if str(fin) in EXCLUDE_DECISIONS:
                continue
            # select minimal article columns
            row = session.exec(select(Article.id, Article.pmid, Article.group_no, Article.year, Article.title_en, Article.title_ja).where(Article.id == aid)).first()
            if not row:
                continue
            _id, pmid, gno, yr, t_en, t_ja = row
            if not pmid:
                continue
            writer.writerow([pmid, label, gno, yr, t_en or "", t_ja or ""])

    output.seek(0)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    if group_no is None:
        filename = f"secondary_pmid_list_{mode}_allgroups_{ts}.csv"
    else:
        filename = f"secondary_pmid_list_{mode}_g{group_no}_{ts}.csv"

    return StreamingResponse(output, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

# --- Export ---
@app.get("/export_disease", response_class=StreamingResponse, name="download_disease")
def export_disease_csv(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    with Session(engine) as session:
        # select only screeningdecision columns that exist to avoid missing-column errors
        cat_cols = ["cat_physical", "cat_brain", "cat_psycho", "cat_drug"]
        has_cols = table_has_columns(session, "screeningdecision", cat_cols)
        sd_fields = [ScreeningDecision.decision, ScreeningDecision.comment, ScreeningDecision.flag_cause, ScreeningDecision.flag_treatment]
        for c in cat_cols:
            if has_cols.get(c, False):
                sd_fields.append(getattr(ScreeningDecision, c))
        # select only article columns that exist to avoid missing-column errors
        article_col_checks = [
            "id", "pmid", "title_en", "title_ja", "direction_gpt", "direction_gemini",
            "condition_list_gpt", "condition_list_gemini", "year",
        ]
        has_article = table_has_columns(session, "article", article_col_checks)
        article_fields = []
        # core fields
        article_fields.append(Article.id)
        article_fields.append(Article.pmid)
        if has_article.get("title_en", False):
            article_fields.append(Article.title_en)
        if has_article.get("title_ja", False):
            article_fields.append(Article.title_ja)
        # optional analysis fields
        if has_article.get("direction_gpt", False):
            article_fields.append(Article.direction_gpt)
        if has_article.get("direction_gemini", False):
            article_fields.append(Article.direction_gemini)
        if has_article.get("condition_list_gpt", False):
            article_fields.append(Article.condition_list_gpt)
        if has_article.get("condition_list_gemini", False):
            article_fields.append(Article.condition_list_gemini)
        if has_article.get("year", False):
            article_fields.append(Article.year)
        user_fields = [User.username]

        stmt = select(*sd_fields, *article_fields, *user_fields).join(Article, Article.id == ScreeningDecision.article_id).join(User, User.id == ScreeningDecision.user_id).order_by(Article.id, User.username)
        try:
            rows = session.exec(stmt).all()
        except Exception:
            # defensive fallback: select minimal article fields only
            article_fields = [Article.id, Article.pmid, Article.title_en, Article.title_ja, Article.year]
            stmt = select(*sd_fields, *article_fields, *user_fields).join(Article, Article.id == ScreeningDecision.article_id).join(User, User.id == ScreeningDecision.user_id).order_by(Article.id, User.username)
            rows = session.exec(stmt).all()

    output = io.StringIO(); output.write("\ufeff"); writer = csv.writer(output)
    writer.writerow(["article_id", "pmid", "title_en", "title_ja", "username", "decision", "comment", 
                     "flag_cause", "flag_treatment", 
                     "cat_physical", "cat_brain", "cat_psycho", "cat_drug", 
                     "direction_gpt", "direction_gemini", "condition_list_gpt", "condition_list_gemini", "year"])

    for row in rows:
        # unpack screeningdecision fields
        idx = 0
        sd_decision = row[idx]; idx += 1
        sd_comment = row[idx]; idx += 1
        sd_flag_cause = int(bool(row[idx])); idx += 1
        sd_flag_treatment = int(bool(row[idx])); idx += 1

        sd_cat_vals = []
        for c in cat_cols:
            if has_cols.get(c, False):
                sd_cat_vals.append(int(bool(row[idx]))); idx += 1
            else:
                sd_cat_vals.append(0)

        # article fields
        art_id = row[idx]; pmid = row[idx+1]
        cur = 2
        title_en = None; title_ja = None; direction_gpt = None; direction_gemini = None; condition_list_gpt = None; condition_list_gemini = None; year = None
        if has_article.get("title_en", False):
            title_en = row[idx + cur - 1]; cur += 1
        if has_article.get("title_ja", False):
            title_ja = row[idx + cur - 1]; cur += 1
        if has_article.get("direction_gpt", False):
            direction_gpt = row[idx + cur - 1]; cur += 1
        if has_article.get("direction_gemini", False):
            direction_gemini = row[idx + cur - 1]; cur += 1
        if has_article.get("condition_list_gpt", False):
            condition_list_gpt = row[idx + cur - 1]; cur += 1
        if has_article.get("condition_list_gemini", False):
            condition_list_gemini = row[idx + cur - 1]; cur += 1
        if has_article.get("year", False):
            year = row[idx + cur - 1]; cur += 1
        idx += cur

        username = row[idx]

        writer.writerow([
            art_id, pmid, title_en, title_ja, username, sd_decision, sd_comment,
            sd_flag_cause, sd_flag_treatment,
            sd_cat_vals[0], sd_cat_vals[1], sd_cat_vals[2], sd_cat_vals[3],
            direction_gpt, direction_gemini, condition_list_gpt, condition_list_gemini, year
        ])
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="apathy_disease_screening_results.csv"'})

@app.get("/export_scale", response_class=StreamingResponse, name="download_scale")
def export_scale_csv(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    with Session(engine) as session:
        rows = session.exec(select(ScaleScreeningDecision, ScaleArticle, User).join(ScaleArticle, ScaleArticle.id == ScaleScreeningDecision.scale_article_id).join(User, User.id == ScaleScreeningDecision.user_id).order_by(ScaleArticle.id, User.username)).all()
    output = io.StringIO(); output.write("\ufeff"); writer = csv.writer(output)
    writer.writerow(["scale_article_id", "pmid", "title_en", "title_ja", "username", "rating", "comment", "gemini_judgement", "gemini_summary_ja", "gemini_reason_ja", "gemini_tools", "year"])
    for sd, art, usr in rows:
        writer.writerow([art.id, art.pmid, art.title_en, art.title_ja, usr.username, sd.rating, sd.comment, art.gemini_judgement, art.gemini_summary_ja, art.gemini_reason_ja, art.gemini_tools, art.year])
    output.seek(0)
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": 'attachment; filename="apathy_scale_screening_results.csv"'})


@app.get("/export_aggregated_disease", response_class=StreamingResponse)
def export_aggregated_disease(request: Request, group_no: Optional[int] = Query(None)):
    """
    Export aggregated disease screening results per article.
    - Does not modify DB.
    - Aggregates per-user `decision` and category flags (`cat_physical`, `cat_brain`, `cat_psycho`, `cat_drug`).
    - Marks category-level conflicts when users disagree on a category flag.
    - Produces a single CSV with per-article aggregation and per-category accepted/hold status.
    """
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    target_group = group_no if group_no is not None else user.group_no
    with Session(engine) as session:
        year_min = get_year_min(session)
        article_ids = get_group_article_ids(session, year_min, target_group)

        # gather decisions: select only the columns that exist to avoid missing-column errors
        cat_cols = ["cat_physical", "cat_brain", "cat_psycho", "cat_drug"]
        has_cols = table_has_columns(session, "screeningdecision", cat_cols)
        fields = [
            ScreeningDecision.article_id,
            ScreeningDecision.user_id,
            ScreeningDecision.decision,
            ScreeningDecision.comment,
        ]
        for c in cat_cols:
            if has_cols.get(c, False):
                fields.append(getattr(ScreeningDecision, c))

        decisions = session.exec(select(*fields).where(ScreeningDecision.article_id.in_(article_ids))).all()

        # map article_id -> list of (username, decision, cat flags, comment)
        art_map = defaultdict(list)
        for row in decisions:
            # row is a tuple aligned with `fields`
            aid = row[0]
            uid = row[1]
            dec = row[2]
            comment = row[3] or ""
            # build category flags safely
            offset = 4
            cat_vals = {}
            for i, c in enumerate(cat_cols):
                if has_cols.get(c, False):
                    val = row[offset + i]
                    cat_vals[c] = int(bool(val))
                else:
                    cat_vals[c] = 0

            if dec is None and not comment and not any(cat_vals.values()):
                # skip empty rows
                continue

            u = session.get(User, uid)
            uname = u.username if u else str(uid)
            art_map[aid].append({
                "username": uname,
                "decision": dec,
                "comment": comment,
                "cat_physical": cat_vals["cat_physical"],
                "cat_brain": cat_vals["cat_brain"],
                "cat_psycho": cat_vals["cat_psycho"],
                "cat_drug": cat_vals["cat_drug"],
            })

        output = io.StringIO(); output.write("\ufeff"); writer = csv.writer(output)
        # header
        writer.writerow([
            "article_id", "pmid", "title_en", "title_ja",
            "aggregated_decision", "decision_counts", "voters_and_decisions", "combined_comment",
            "cat_physical_votes", "cat_physical_final", "cat_physical_conflict",
            "cat_brain_votes", "cat_brain_final", "cat_brain_conflict",
            "cat_psycho_votes", "cat_psycho_final", "cat_psycho_conflict",
            "cat_drug_votes", "cat_drug_final", "cat_drug_conflict",
            "year"
        ])

        for aid in sorted(article_ids):
            art = get_article_safe(session, aid)
            rows = art_map.get(aid, [])

            # aggregated decision logic: list counts for 0/1/2, and majority (max of votes) as in existing export
            counts = {0:0, 1:0, 2:0}
            voters = []
            combined_comments = []
            for r in rows:
                dec = r.get("decision")
                if dec is not None:
                    try: counts[int(dec)] += 1
                    except: pass
                voters.append(f"{r['username']}:{r.get('decision')}")
                if r.get('comment'):
                    combined_comments.append(f"{r['username']}:{r.get('comment')}")

            agg_decision = None
            if any(counts.values()):
                # choose highest voted level by count; fallback to max vote value
                agg_decision = max(counts, key=lambda k: (counts[k], k))

            # categories: collect votes per category and detect conflict (disagreement among voters)
            def analyze_cat(cat_name):
                votes = [r.get(cat_name) for r in rows if r.get(cat_name) is not None]
                votes = [int(v) for v in votes]
                votes_str = "+".join(str(v) for v in votes) if votes else ""
                final = 1 if votes and max(votes) >= 1 else 0
                conflict = False
                if votes:
                    has1 = any(v == 1 for v in votes)
                    has0 = any(v == 0 for v in votes)
                    if has1 and has0:
                        conflict = True
                return votes_str, final, int(conflict)

            p_votes, p_final, p_conf = analyze_cat("cat_physical")
            b_votes, b_final, b_conf = analyze_cat("cat_brain")
            y_votes, y_final, y_conf = analyze_cat("cat_psycho")
            d_votes, d_final, d_conf = analyze_cat("cat_drug")

            writer.writerow([
                art.id, art.pmid, art.title_en, art.title_ja,
                agg_decision, f"0:{counts[0]}|1:{counts[1]}|2:{counts[2]}", ";".join(voters), ";".join(combined_comments),
                p_votes, p_final, p_conf,
                b_votes, b_final, b_conf,
                y_votes, y_final, y_conf,
                d_votes, d_final, d_conf,
                art.year
            ])

    output.seek(0)
    filename = f"aggregated_disease_group{target_group}.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/export_category_lists", response_class=StreamingResponse)
def export_category_lists(request: Request, group_no: Optional[int] = Query(None)):
    """
    Export per-category lists of articles marked as accepted (final=1) or hold (final=0) for secondary screening.
    Uses the same rule as `/export_aggregated_disease` for category finalization (any vote of 1 -> accepted).
    """
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    # NOTE: category list exports should include all groups (ignore per-group slicing)
    with Session(engine) as session:
        year_min = get_year_min(session)
        # collect all article ids respecting year_min
        all_rows = list(session.exec(select(Article.id, Article.year)).all())
        if year_min is not None:
            filtered = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
            rows = filtered if filtered else all_rows
        else:
            rows = all_rows
        # rows are tuples (id, year)
        article_ids = [r[0] for r in rows]

        # collect screening decisions: select only category columns that exist
        cat_cols = ["cat_physical", "cat_brain", "cat_psycho", "cat_drug"]
        has_cols = table_has_columns(session, "screeningdecision", cat_cols)
        fields = [ScreeningDecision.article_id]
        for c in cat_cols:
            if has_cols.get(c, False):
                fields.append(getattr(ScreeningDecision, c))
        decisions = session.exec(select(*fields).where(ScreeningDecision.article_id.in_(article_ids))).all()
        art_map = defaultdict(list)
        for row in decisions:
            aid = row[0]
            d = {}
            offset = 1
            for i, c in enumerate(cat_cols):
                if has_cols.get(c, False):
                    d[c] = row[offset + i]
            art_map[aid].append(d)

        # produce CSV with category sections separated by header rows
        output = io.StringIO(); output.write("\ufeff"); writer = csv.writer(output)

        def write_section(cat_label, extractor):
            writer.writerow([cat_label])
            writer.writerow(["article_id", "pmid", "title", "status", "votes_summary"])
            for aid in sorted(article_ids):
                art = get_article_safe(session, aid)
                rows = art_map.get(aid, [])
                votes = [int(r.get(extractor)) for r in rows if r.get(extractor) is not None]
                if votes and max(votes) >= 1:
                    status = "accepted"
                else:
                    status = "hold"
                votes_summary = "+".join(str(v) for v in votes) if votes else ""
                writer.writerow([art.id, art.pmid, art.title_ja or art.title_en, status, votes_summary])
            writer.writerow([])

        write_section("cat_physical", "cat_physical")
        write_section("cat_brain", "cat_brain")
        write_section("cat_psycho", "cat_psycho")
        write_section("cat_drug", "cat_drug")

    output.seek(0)
    filename = f"category_lists_allgroups.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _export_category_csv(session: Session, article_ids: List[int], cat_attr: str):
    """Helper: return StringIO CSV for a single category where
    - at least one rater selected the category, and
    - the aggregated decision for the article is採用 (any user decision >= 1)
    """
    output = io.StringIO(); output.write("\ufeff"); writer = csv.writer(output)
    writer.writerow(["article_id", "pmid", "title_en", "title_ja", "aggregated_decision", "category_votes"])

    # select only necessary columns (decision and requested category) to avoid missing-column errors
    has_cat = table_has_columns(session, "screeningdecision", [cat_attr]).get(cat_attr, False)
    fields = [ScreeningDecision.article_id, ScreeningDecision.decision]
    if has_cat:
        fields.append(getattr(ScreeningDecision, cat_attr))
    decisions = session.exec(select(*fields).where(ScreeningDecision.article_id.in_(article_ids))).all()
    art_map = defaultdict(list)
    for row in decisions:
        aid = row[0]
        dec = row[1]
        cat_val = None
        if has_cat:
            cat_val = row[2]
        art_map[aid].append({"decision": dec, cat_attr: cat_val})

    for aid in sorted(article_ids):
        art = get_article_safe(session, aid)
        rows = art_map.get(aid, [])
        # aggregated decision checks
        dec_votes = [int(d["decision"]) for d in rows if d.get("decision") is not None]
        if not dec_votes or max(dec_votes) < 1:
            continue

        # category votes
        cat_votes = [int(d.get(cat_attr)) for d in rows if d.get(cat_attr) is not None]
        if not cat_votes or max(cat_votes) < 1:
            continue

        if not art:
            continue
        votes_summary = "+".join(str(v) for v in cat_votes)
        agg_decision = max(dec_votes) if dec_votes else ""
        writer.writerow([art.id, art.pmid, art.title_en, art.title_ja, agg_decision, votes_summary])

    output.seek(0)
    return output


@app.get("/export_category_physical", response_class=StreamingResponse)
def export_category_physical(request: Request, group_no: Optional[int] = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    # Export should include all groups regardless of `group_no`
    with Session(engine) as session:
        year_min = get_year_min(session)
        all_rows = list(session.exec(select(Article.id, Article.year)).all())
        if year_min is not None:
            rows = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
            if not rows:
                rows = all_rows
        else:
            rows = all_rows
        article_ids = [r[0] for r in rows]
        output = _export_category_csv(session, article_ids, "cat_physical")
    filename = f"category_physical_allgroups.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/export_category_brain", response_class=StreamingResponse)
def export_category_brain(request: Request, group_no: Optional[int] = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    # Export should include all groups regardless of `group_no`
    with Session(engine) as session:
        year_min = get_year_min(session)
        all_rows = list(session.exec(select(Article.id, Article.year)).all())
        if year_min is not None:
            rows = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
            if not rows:
                rows = all_rows
        else:
            rows = all_rows
        article_ids = [r[0] for r in rows]
        output = _export_category_csv(session, article_ids, "cat_brain")
    filename = f"category_brain_allgroups.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/export_category_psycho", response_class=StreamingResponse)
def export_category_psycho(request: Request, group_no: Optional[int] = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    # Export should include all groups regardless of `group_no`
    with Session(engine) as session:
        year_min = get_year_min(session)
        all_rows = list(session.exec(select(Article.id, Article.year)).all())
        if year_min is not None:
            rows = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
            if not rows:
                rows = all_rows
        else:
            rows = all_rows
        article_ids = [r[0] for r in rows]
        output = _export_category_csv(session, article_ids, "cat_psycho")
    filename = f"category_psycho_allgroups.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/export_category_drug", response_class=StreamingResponse)
def export_category_drug(request: Request, group_no: Optional[int] = Query(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login")
    # Export should include all groups regardless of `group_no`
    with Session(engine) as session:
        year_min = get_year_min(session)
        all_rows = list(session.exec(select(Article.id, Article.year)).all())
        if year_min is not None:
            rows = [r for r in all_rows if (r[1] is not None and r[1] >= year_min)]
            if not rows:
                rows = all_rows
        else:
            rows = all_rows
        article_ids = [r[0] for r in rows]
        output = _export_category_csv(session, article_ids, "cat_drug")
    filename = f"category_drug_allgroups.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# =========================================================
# Secondary (二次) screening routes
# =========================================================
@app.get("/secondary", response_class=HTMLResponse)
def secondary_index(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    groups = ["physical", "brain", "psycho", "drug"]
    stats = {}
    candidates_by_group = {}  # New: store candidates with review status for display
    
    with Session(engine) as session:
        for g in groups:
            col = getattr(SecondaryArticle, f"is_{g}")
            total = session.exec(select(func.count(SecondaryArticle.id)).where(col == True)).one() if session.exec(select(func.count(SecondaryArticle.id)).where(col == True)).one() is not None else 0

            if user.is_admin:
                # Admin: pending = articles in this group that admin has not completed (no review or pending)
                pmid_rows = session.exec(select(SecondaryArticle.pmid).where(col == True).order_by(SecondaryArticle.pmid)).all()
                pending_count = 0
                included_count = 0
                excluded_count = 0
                completed_count = 0
                
                candidates_data = []
                for pr in pmid_rows:
                    pmid = pr
                    rev = session.exec(select(SecondaryReview).where((SecondaryReview.group == g) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.pmid == pmid))).first()
                    
                    if not rev:
                        pending_count += 1
                        status = "pending"
                        is_completed = False
                    else:
                        is_completed = rev.completed_at is not None
                        if rev.decision == 'pending':
                            pending_count += 1
                            status = "pending"
                        elif rev.decision == 'include':
                            included_count += 1
                            status = "include"
                        elif rev.decision == 'exclude':
                            excluded_count += 1
                            status = "exclude"
                        else:
                            status = rev.decision
                        
                        if is_completed:
                            completed_count += 1
                    
                    candidates_data.append({
                        "pmid": pmid,
                        "decision": rev.decision if rev else None,
                        "status": status,
                        "completed_at": rev.completed_at if rev else None,
                        "is_completed": is_completed
                    })
                
                stats[g] = {"total": total or 0, "pending": pending_count, "included": included_count, "excluded": excluded_count, "completed": completed_count}
                candidates_by_group[g] = candidates_data
            else:
                pending = session.exec(select(func.count(SecondaryReview.id)).where((SecondaryReview.group == g) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.decision == "pending"))).one()
                included = session.exec(select(func.count(SecondaryReview.id)).where((SecondaryReview.group == g) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.decision == "include"))).one()
                excluded = session.exec(select(func.count(SecondaryReview.id)).where((SecondaryReview.group == g) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.decision == "exclude"))).one()
                completed = session.exec(select(func.count(SecondaryReview.id)).where((SecondaryReview.group == g) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.completed_at != None))).one()
                
                stats[g] = {"total": total or 0, "pending": pending or 0, "included": included or 0, "excluded": excluded or 0, "completed": completed or 0}
                
                # For non-admins, also get candidate list for display
                pmid_rows = session.exec(select(SecondaryArticle.pmid).where(col == True).order_by(SecondaryArticle.pmid)).all()
                candidates_data = []
                for pmid in pmid_rows:
                    rev = session.exec(select(SecondaryReview).where((SecondaryReview.group == g) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.pmid == pmid))).first()
                    if rev:
                        status = "pending" if rev.decision == "pending" else rev.decision
                        candidates_data.append({
                            "pmid": pmid,
                            "decision": rev.decision,
                            "status": status,
                            "completed_at": rev.completed_at,
                            "is_completed": rev.completed_at is not None
                        })
                
                candidates_by_group[g] = candidates_data

    return templates.TemplateResponse("secondary_index.html", {
        "request": request, 
        "username": user.username, 
        "group_no": user.group_no, 
        "stats": stats,
        "candidates_by_group": candidates_by_group,
        "current_page": "secondary"
    })


@app.get("/secondary/{group}/next")
def secondary_next(request: Request, group: str):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    with Session(engine) as session:
        if user.is_admin:
            # find first SecondaryArticle in this group that admin hasn't completed (no review) or has pending
            col = getattr(SecondaryArticle, f"is_{group}")
            rows = session.exec(select(SecondaryArticle.pmid).where(col == True).order_by(SecondaryArticle.pmid)).all()
            for r in rows:
                pmid = r
                rev = session.exec(select(SecondaryReview).where((SecondaryReview.group == group) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.pmid == pmid))).first()
                if not rev or (rev and rev.decision == 'pending'):
                    return RedirectResponse(f"/secondary/{group}/{pmid}", 303)
        else:
            nxt = session.exec(select(SecondaryReview).where((SecondaryReview.group == group) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.decision == "pending")).order_by(SecondaryReview.pmid)).first()
            if nxt:
                return RedirectResponse(f"/secondary/{group}/{nxt.pmid}", 303)
        
        # Get all completed items for this group to show in empty page
        completed_reviews = session.exec(
            select(SecondaryReview).where(
                (SecondaryReview.group == group) & 
                (SecondaryReview.reviewer_id == user.id)
            ).order_by(SecondaryReview.pmid)
        ).all()
        
    return templates.TemplateResponse("secondary_empty.html", {
        "request": request, 
        "username": user.username, 
        "group": group, 
        "current_page": "secondary",
        "completed_reviews": completed_reviews
    })


@app.get("/secondary/{group}/{pmid}", response_class=HTMLResponse)
def secondary_review_page(request: Request, group: str, pmid: int):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    # Log which DB and request context we're using to aid debugging
    try:
        logger.info("secondary view: pmid=%s group=%s user=%s DATABASE_URL=%s", pmid, group, user.username, engine.url)
    except Exception:
        # best-effort logging
        print(f"secondary view: pmid={pmid} group={group} user={user.username} DATABASE_URL={engine.url}")

    auto_error = None
    with Session(engine) as session:
        article = session.exec(select(Article).where(Article.pmid == pmid)).first()
        secondary = session.exec(select(SecondaryArticle).where(SecondaryArticle.pmid == pmid)).first()
        # Load auto extraction table defensively: if table missing, do not raise 500
        try:
            auto_obj = session.exec(select(SecondaryAutoExtraction).where(SecondaryAutoExtraction.pmid == pmid)).first()
        except OperationalError as e:
            logger.error("SecondaryAutoExtraction query failed: %s", e, exc_info=True)
            auto_obj = None
            auto_error = "Gemini下書きテーブルが未作成のため表示できません（管理者に連絡してください）"
        except Exception:
            # re-raise unexpected exceptions so they appear in logs and fail loud
            logger.exception("Unexpected error loading SecondaryAutoExtraction")
            raise

        review = session.exec(select(SecondaryReview).where((SecondaryReview.pmid == pmid) & (SecondaryReview.group == group) & (SecondaryReview.reviewer_id == user.id))).first()
        if not review:
            if user.is_admin:
                # create a pending review record for admin on-the-fly
                review = SecondaryReview(pmid=pmid, group=group, reviewer_id=user.id, decision="pending")
                session.add(review); session.commit()
                review = session.exec(select(SecondaryReview).where((SecondaryReview.pmid == pmid) & (SecondaryReview.group == group) & (SecondaryReview.reviewer_id == user.id))).first()
            else:
                # non-admins should not view unassigned items
                return templates.TemplateResponse("secondary_empty.html", {"request": request, "username": user.username, "group": group, "current_page": "secondary"})

        pdf_available = False
        pdf_dir = os.getenv("SECONDARY_PDF_DIR", DEFAULT_SECONDARY_PDF_DIR)
        if pdf_dir:
            pdf_path = Path(pdf_dir) / f"{pmid}.pdf"
            pdf_available = pdf_path.exists()
        # Serialize ORM objects to plain dicts for template safety (avoid DetachedInstanceError)
        # compute progress for this reviewer within the group
        try:
            id_rows = session.exec(select(SecondaryArticle.pmid).where(getattr(SecondaryArticle, f"is_{group}") == True).order_by(SecondaryArticle.pmid)).all()
            id_list = [int(r) for r in id_rows] if id_rows else []
        except Exception:
            id_list = []

        progress_total = len(id_list)
        if progress_total:
            done_count = session.exec(select(func.count(SecondaryReview.id)).where(
                (SecondaryReview.group == group) & (SecondaryReview.reviewer_id == user.id) & (SecondaryReview.pmid.in_(id_list)) & (SecondaryReview.decision != 'pending')
            )).one()
        else:
            done_count = 0

        try:
            current_index = id_list.index(pmid) + 1 if pmid in id_list else None
        except Exception:
            current_index = None

        article = _serialize_article(article)
        secondary = _serialize_secondary(secondary)
        auto = _serialize_auto(auto_obj)
        review = _serialize_review(review)
    return templates.TemplateResponse("secondary_review.html", {
        "request": request,
        "username": user.username,
        "group": group,
        "pmid": pmid,
        "article": article,
        "secondary": secondary,
        "auto": auto,
        "auto_error": auto_error,
        "review": review,
        "progress_done": int(done_count) if isinstance(done_count, (int,)) or hasattr(done_count, '__int__') else int(done_count[0]) if (done_count and isinstance(done_count, tuple)) else int(done_count),
        "progress_total": progress_total,
        "current_index": current_index,
        "pdf_available": pdf_available,
        "current_page": "secondary"
    })


@app.post("/secondary/{group}/{pmid}/save")
def secondary_save(request: Request, group: str, pmid: int,
                   decision: str = Form("pending"), final_citation: str = Form(""), final_apathy_terms: str = Form(""),
                   final_target_condition: str = Form(""), final_population_n: str = Form(""), final_prevalence: str = Form(""), final_intervention: str = Form(""),
                   comment: str = Form(""), action: str = Form("save"), nav: str = Form(None), jump_index: str | None = Form(None)):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    with Session(engine) as session:
        review = session.exec(select(SecondaryReview).where((SecondaryReview.pmid == pmid) & (SecondaryReview.group == group) & (SecondaryReview.reviewer_id == user.id))).first()
        if not review:
            if user.is_admin:
                review = SecondaryReview(pmid=pmid, group=group, reviewer_id=user.id)
            else:
                # non-admin cannot save for unassigned item
                return RedirectResponse(f"/secondary/{group}/next", 303)
        
        # If user clicked the "完了として保存" button, mark as completed
        if action == 'complete':
            review.completed_at = datetime.utcnow().isoformat()
        
        # If user clicked the explicit "除外して次へ" button, force decision to exclude
        if action == 'exclude_next':
            review.decision = 'exclude'
        else:
            review.decision = decision
        review.final_citation = final_citation
        review.final_apathy_terms = final_apathy_terms
        review.final_target_condition = final_target_condition
        # persist both variants if model has them for compatibility
        review.final_population_n = final_population_n
        review.final_prevalence = final_prevalence
        review.final_intervention = final_intervention
        review.comment = comment
        review.updated_at = datetime.utcnow().isoformat()
        session.add(review); session.commit()
        # compute id list for navigation actions
        try:
            id_rows = session.exec(select(SecondaryArticle.pmid).where(getattr(SecondaryArticle, f"is_{group}") == True).order_by(SecondaryArticle.pmid)).all()
            id_list = [int(r) for r in id_rows] if id_rows else []
        except Exception:
            id_list = []

    # Navigation handling
    if action == 'complete':
        # After marking complete, stay on same page to confirm
        return RedirectResponse(f"/secondary/{group}/{pmid}", 303)
    
    if action == 'exclude_next':
        return RedirectResponse(f"/secondary/{group}/next", 303)

    if nav == 'prev' and id_list:
        try:
            idx = id_list.index(pmid)
            target = id_list[idx - 1] if idx > 0 else id_list[0]
            return RedirectResponse(f"/secondary/{group}/{target}", 303)
        except ValueError:
            return RedirectResponse(f"/secondary/{group}/next", 303)

    if nav == 'jump' and id_list and jump_index:
        try:
            j = int(jump_index)
            j = max(1, min(j, len(id_list)))
            target = id_list[j - 1]
            return RedirectResponse(f"/secondary/{group}/{target}", 303)
        except Exception:
            return RedirectResponse(f"/secondary/{group}/{pmid}", 303)

    if nav == 'next' or action == 'save_next':
        return RedirectResponse(f"/secondary/{group}/next", 303)

    # default: return to same page
    return RedirectResponse(f"/secondary/{group}/{pmid}", 303)


@app.get("/secondary/pdf/{pmid}")
def secondary_pdf(pmid: int):
    pdf_dir = os.getenv("SECONDARY_PDF_DIR", DEFAULT_SECONDARY_PDF_DIR)
    if not pdf_dir:
        return HTMLResponse("PDF directory not configured", status_code=404)
    path = Path(pdf_dir) / f"{pmid}.pdf"
    if not path.exists():
        return HTMLResponse("PDF not found", status_code=404)
    return FileResponse(path, media_type="application/pdf", headers={"Content-Disposition": f'inline; filename="{pmid}.pdf"'})


@app.get("/pdf/{pmid}")
def pdf_redirect(pmid: int, request: Request):
    user = get_current_user(request)
    if not user:
        return RedirectResponse("/login", status_code=302)
    try:
        pmid_i = int(pmid)
    except Exception:
        raise HTTPException(status_code=400, detail="invalid pmid")
    if pmid_i <= 0:
        raise HTTPException(status_code=400, detail="bad pmid")
    url = build_pdf_url(pmid_i)
    return RedirectResponse(url, status_code=302)


@app.get("/secondary/{group}/export")
def secondary_group_export(request: Request, group: str, format: str = Query("csv")):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    
    # Fetch all reviews for this group with auto-extraction and user info
    with Session(engine) as session:
        rows = session.exec(
            select(SecondaryReview, SecondaryAutoExtraction, User, Article)
            .join(User, User.id == SecondaryReview.reviewer_id)
            .outerjoin(SecondaryAutoExtraction, SecondaryAutoExtraction.pmid == SecondaryReview.pmid)
            .outerjoin(Article, Article.pmid == SecondaryReview.pmid)
            .where(SecondaryReview.group == group)
        ).all()
    
    if format.lower() == "xlsx":
        return _export_secondary_xlsx(rows, group)
    else:
        return _export_secondary_csv(rows, group)


def _export_secondary_csv(rows, group: str):
    """Export secondary reviews as CSV (vertical format: 1 row = 1 PMID × reviewer)"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header: aligned with template requirements
    writer.writerow([
        "pmid",
        "reviewer",
        "decision",
        "target_condition",
        "apathy_terms",
        "population_n",
        "prevalence",
        "intervention",
        "comment",
        "auto_target_condition",
        "auto_apathy_terms",
        "auto_population_n",
        "auto_prevalence",
        "auto_intervention"
    ])
    
    # Data rows
    for rev, auto, usr, article in rows:
        writer.writerow([
            rev.pmid,
            usr.username if usr else str(rev.reviewer_id),
            rev.decision or "",
            rev.final_target_condition or "",
            rev.final_apathy_terms or "",
            rev.final_population_n or "",
            rev.final_prevalence or "",
            rev.final_intervention or "",
            rev.comment or "",
            auto.auto_target_condition if auto else "",
            auto.auto_apathy_terms if auto else "",
            auto.auto_population_N if auto else "",
            auto.auto_prevalence if auto else "",
            auto.auto_intervention if auto else ""
        ])
    
    output.seek(0)
    filename = f"secondary_group_{group}_export.csv"
    return StreamingResponse(output, media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})


def _export_secondary_xlsx(rows, group: str):
    """Export secondary reviews as XLSX using openpyxl (vertical format)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HTMLResponse("openpyxl not installed. Please install it to use XLSX export.", status_code=500)
    
    wb = Workbook()
    ws = wb.active
    ws.title = group
    
    # Header row with styling
    header_row = [
        "PMID",
        "Reviewer",
        "Decision",
        "Target Condition",
        "Apathy Terms",
        "Population N",
        "Prevalence",
        "Intervention",
        "Comment",
        "Auto Target Condition",
        "Auto Apathy Terms",
        "Auto Population N",
        "Auto Prevalence",
        "Auto Intervention"
    ]
    ws.append(header_row)
    
    # Style header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data rows
    for rev, auto, usr, article in rows:
        ws.append([
            rev.pmid,
            usr.username if usr else str(rev.reviewer_id),
            rev.decision or "",
            rev.final_target_condition or "",
            rev.final_apathy_terms or "",
            rev.final_population_n or "",
            rev.final_prevalence or "",
            rev.final_intervention or "",
            rev.comment or "",
            auto.auto_target_condition if auto else "",
            auto.auto_apathy_terms if auto else "",
            auto.auto_population_N if auto else "",
            auto.auto_prevalence if auto else "",
            auto.auto_intervention if auto else ""
        ])
    
    # Column widths
    ws.column_dimensions["A"].width = 12  # PMID
    ws.column_dimensions["B"].width = 15  # Reviewer
    ws.column_dimensions["C"].width = 12  # Decision
    ws.column_dimensions["D"].width = 25  # Target Condition
    ws.column_dimensions["E"].width = 25  # Apathy Terms
    ws.column_dimensions["F"].width = 20  # Population N
    ws.column_dimensions["G"].width = 20  # Prevalence
    ws.column_dimensions["H"].width = 25  # Intervention
    ws.column_dimensions["I"].width = 20  # Comment
    for col in ["J", "K", "L", "M", "N"]:
        ws.column_dimensions[col].width = 20
    
    # Output to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"secondary_group_{group}_export.xlsx"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/secondary/conditions/summary")
def secondary_conditions_summary(request: Request):
    user = get_current_user(request)
    if not user: return RedirectResponse("/login", 303)
    from collections import defaultdict
    cnt = defaultdict(list)
    with Session(engine) as session:
        rows = session.exec(select(SecondaryReview).where(SecondaryReview.final_target_condition.is_not(None))).all()
        for r in rows:
            key = (r.final_target_condition or "").strip()
            if not key: continue
            cnt[key].append(r.pmid)
    out = {k: {"count": len(v), "pmids": v} for k, v in cnt.items()}
    return out